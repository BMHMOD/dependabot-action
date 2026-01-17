import openpyxl
from tkinter import Tk, filedialog, messagebox
import os
import time

# This script requires the 'openpyxl' library.
# You can install it by running: pip install openpyxl

def select_file(title):
    """Opens a dialog to select an Excel file and returns its path."""
    root = Tk()
    root.withdraw()  # Hide the main window
    root.attributes('-topmost', True)  # Bring to front
    filepath = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel Files", "*.xlsx;*.xlsm"), ("All files", "*.*")]
    )
    root.destroy()
    return filepath

def main():
    """Main function to process the yard files from the stock file."""
    try:
        print("--- Container Tracker: Yard Fill Tool (Enhanced) ---")
        
        # 1. Select the files
        stock_path = select_file("الرجاء اختيار ملف STOCK")
        if not stock_path:
            print("تم الإلغاء. لم يتم اختيار ملف STOCK.")
            return

        internal_path = select_file("الرجاء اختيار ملف Internal Yard")
        if not internal_path:
            print("تم الإلغاء. لم يتم اختيار ملف Internal Yard.")
            return

        external_path = select_file("الرجاء اختيار ملف External Yard")
        if not external_path:
            print("تم الإلغاء. لم يتم اختيار ملف External Yard.")
            return

        print("جاري فتح الملفات...")
        start_time = time.time()

        # 2. Open the workbooks
        # data_only=True helps read values computed by formulas if needed, 
        # but here we need to read raw data from STOCK and write to others.
        wb_stock = openpyxl.load_workbook(stock_path, data_only=True)
        ws_stock = wb_stock.active

        wb_internal = openpyxl.load_workbook(internal_path)
        ws_internal = wb_internal.active

        wb_external = openpyxl.load_workbook(external_path)
        ws_external = wb_external.active

        print("جاري مسح البيانات القديمة...")

        # 3. Clear old data from yard files
        # Clear Internal Yard (Columns C:G = 3:7)
        # Clearing slightly more rows to be safe (up to row 100 as per VBA enhanced)
        for row in ws_internal.iter_rows(min_row=6, max_row=100, min_col=3, max_col=7):
            for cell in row:
                cell.value = None

        # Clear External Yard (Columns C:F = 3:6)
        # Clearing up to row 30 as per VBA enhanced
        for row in ws_external.iter_rows(min_row=6, max_row=30, min_col=3, max_col=6):
            for cell in row:
                cell.value = None

        # 4. Define mappings
        # Format: "BlockName": StartRow
        block_rows = {
            "M": 6,
            "A": 9,
            "B": 12,
            "C": 15,
            "D": 18,
            "H": 21,
            "F": 24,
            "Y777": 29,
            "S22": 35,
            "S003": 38,
            "S666": 41,
            "INSP": 44,
            "S002": 47,  # Added
            "S03": 50,   # Added
            "S333": 53   # Added
        }
        
        # Format: "YardName": (StartRow, "Area1|Area2|Block1|...")
        external_yards = {
            "التجارية": (6, "|S444|S068|S032|"),
            "المفروزة": (8, "|S900|RORO1|BR|"),
            "68": (10, "|S600|S700|"),
            "Other1": (12, "|RAIL|SCALE|"),
            "Other2": (14, "|XRAY|RORO5|")
        }

        # 5. Process Data
        print("جاري معالجة البيانات من ملف STOCK...")
        
        max_row = ws_stock.max_row
        processed_count = 0
        
        for i in range(2, max_row + 1):
            if i % 500 == 0:
                print(f"تم معالجة {i} من {max_row} صف...")
            
            # Read values (using safe string conversion)
            # Column P (16): Mode, G (7): Block, M (13): FE, J (10): Cntr Len, F (6): Area
            
            mode_val = str(ws_stock.cell(row=i, column=16).value or "").strip().upper()
            block_val = str(ws_stock.cell(row=i, column=7).value or "").strip().upper()
            fe_val = str(ws_stock.cell(row=i, column=13).value or "").strip().upper()
            cntr_len = str(ws_stock.cell(row=i, column=10).value or "").strip()
            area_val = str(ws_stock.cell(row=i, column=6).value or "").strip().upper()

            processed_count += 1
            
            # --- Process Internal Yard ---
            if block_val in block_rows:
                base_row = block_rows[block_val]
                target_row = 0
                
                if mode_val == "IMPORT":
                    target_row = base_row
                elif mode_val == "EXPORT":
                    target_row = base_row + 1
                elif mode_val in ["STORAGE", "TRANSSHIPMENT"]:
                    target_row = base_row + 2
                
                if target_row > 0:
                    target_col = 0
                    if cntr_len == "20" and fe_val == "F": target_col = 3   # C
                    elif cntr_len == "40" and fe_val == "F": target_col = 4 # D
                    elif cntr_len == "20" and fe_val == "E": target_col = 5 # E
                    elif cntr_len == "40" and fe_val == "E": target_col = 6 # F
                    elif cntr_len == "45": target_col = 7                   # G (Internal Only)
                    
                    if target_col > 0:
                        current_val = ws_internal.cell(row=target_row, column=target_col).value
                        if current_val is None: current_val = 0
                        ws_internal.cell(row=target_row, column=target_col).value = current_val + 1

            # --- Process External Yard ---
            # Check through all external yards defined
            # Logic: Match if (Area OR Block) is inside the pipe-delimited string
            
            found_yard = False
            for yard_name, (start_row, yard_areas) in external_yards.items():
                if found_yard: break # Optimization: One container usually belongs to one place
                
                # Check for pipe-delimited exact match
                # e.g. "|S444|S068|" contains "|S444|"
                check_area = f"|{area_val}|"
                check_block = f"|{block_val}|"
                
                if (check_area in yard_areas) or (check_block in yard_areas):
                    ext_target_row = 0
                    if mode_val == "IMPORT":
                        ext_target_row = start_row
                    elif mode_val == "EXPORT":
                        ext_target_row = start_row + 1
                        
                    if ext_target_row > 0:
                        ext_target_col = 0
                        if cntr_len == "20" and fe_val == "F": ext_target_col = 3   # C
                        elif cntr_len == "40" and fe_val == "F": ext_target_col = 4 # D
                        elif cntr_len == "20" and fe_val == "E": ext_target_col = 5 # E
                        elif cntr_len == "40" and fe_val == "E": ext_target_col = 6 # F
                        
                        if ext_target_col > 0:
                            current_val = ws_external.cell(row=ext_target_row, column=ext_target_col).value
                            if current_val is None: current_val = 0
                            ws_external.cell(row=ext_target_row, column=ext_target_col).value = current_val + 1
                    
                    found_yard = True

        # 6. Save the files
        print("جاري حفظ الملفات...")
        wb_internal.save(internal_path)
        wb_external.save(external_path)

        elapsed = time.time() - start_time
        msg = f"تم الانتهاء بنجاح!\nعدد الصفوف المعالجة: {processed_count}\nالوقت: {elapsed:.2f} ثانية"
        
        print(msg)
        
        # Show success message box
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        messagebox.showinfo("اكتمل", msg)
        root.destroy()

    except Exception as e:
        err_msg = f"حدث خطأ أثناء التشغيل:\n{str(e)}"
        print(err_msg)
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        messagebox.showerror("خطأ", err_msg)
        root.destroy()

if __name__ == "__main__":
    main()
